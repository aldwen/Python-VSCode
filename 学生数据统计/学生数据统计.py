import openpyxl
import math
filename=r'd:\aldwen\coding\python\学生数据统计\input\2019source.xlsx'
openfile=openpyxl.load_workbook(filename)

sheetnames=openfile.get_sheet_names()
firstSheet=openfile.get_sheet_by_name(sheetnames[0])
students=[]
b总人数=firstSheet.max_row
print(b总人数)
for row in firstSheet.rows:
    a=[]
    for cell in row:
        a.append(cell.value)
    students.append(a)   
b汇总={} 
students=sorted(students,key=(lambda x:x[5]),reverse=True)
bA等级级分数线=students[math.ceil(0.2*b总人数)-1][5]
bB等级级分数线=students[math.ceil(0.4*b总人数)-1][5]
bC等级级分数线=students[math.ceil(0.6*b总人数)-1][5]
bD等级级分数线=students[math.ceil(0.8*b总人数)-1][5]

b=[bA等级级分数线,bB等级级分数线,bC等级级分数线,bD等级级分数线,0]

for astu in students:
    if astu[3] not in b汇总:
        b汇总[astu[3]]={"A":0,"B":0,"C":0,"D":0,"E":0,}

    if astu[5]>=bA等级级分数线:
        b汇总[astu[3]]["A"]+=1
    elif astu[5]>=bB等级级分数线:
        b汇总[astu[3]]["B"]+=1
    elif astu[5]>=bC等级级分数线:
        b汇总[astu[3]]["C"]+=1
    elif astu[5]>=bD等级级分数线:
        b汇总[astu[3]]["D"]+=1
    else:
        b汇总[astu[3]]["E"]+=1

    if astu[2] not in b汇总[astu[3]]:
        b汇总[astu[3]][astu[2]]={"A":0,"B":0,"C":0,"D":0,"E":0,}
    if astu[5]>=bA等级级分数线:
        b汇总[astu[3]][astu[2]]["A"]+=1
    if astu[5]>=bB等级级分数线:
        b汇总[astu[3]][astu[2]]["B"]+=1        
    if astu[5]>=bC等级级分数线:
        b汇总[astu[3]][astu[2]]["C"]+=1
    if astu[5]>=bD等级级分数线:
        b汇总[astu[3]][astu[2]]["D"]+=1   
    else:
        b汇总[astu[3]][astu[2]]["E"]+=1        


print(b汇总)